
import streamlit as st
import pandas as pd
from PIL import Image, ImageDraw, ImageFont, ExifTags
import io
import datetime
import socket
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as ExcelImage
from pyngrok import ngrok, conf
from pyngrok.exception import PyngrokNgrokError

# ==========================================
# 設定
# ==========================================
st.set_page_config(page_title="工事写真台帳作成アプリ", layout="wide")
st.title("📷 工事写真台帳作成アプリ")
st.write("工事前の写真をアップロードして、工務店への見積依頼用エクセルを作成します。")

# ==========================================
# サイドバー設定
# ==========================================
st.sidebar.header("設定")

# 日付設定
date_mode = st.sidebar.radio(
    "日付の印字",
    ("指定日を入れる", "写真の撮影日(Exif)", "印字しない")
)

date_text_fixed = ""
if date_mode == "指定日を入れる":
    date_input = st.sidebar.date_input("撮影日指定", datetime.date.today())
    date_text_fixed = date_input.strftime('%Y.%m.%d')
elif date_mode == "写真の撮影日(Exif)":
    st.sidebar.info("写真に撮影日情報(Exif)がない場合は印字されません。")

# スマホアクセスの案内
st.sidebar.markdown("---")
st.sidebar.subheader("スマホからアクセス")

# 1. ローカルLANアクセス
try:
    hostname = socket.gethostname()
    ip_address = socket.gethostbyname(hostname)
    st.sidebar.write(f"**Wi-Fi内アクセス:**")
    st.sidebar.code(f"http://{ip_address}:8501")
except:
    pass

st.sidebar.markdown("---")

# 2. 外部公開（ngrok）
st.sidebar.subheader("外部公開 (ngrok)")
st.sidebar.info("Wi-Fi外（4G/5Gなど）からアクセスしたい場合はこちら")

# キャッシュリソースとしてトンネル管理
@st.cache_resource
def get_ngrok_tunnel(auth_token=None):
    if auth_token:
        # トークン設定
        ngrok.set_auth_token(auth_token)
    
    try:
        # 既存のトンネルを確認
        tunnels = ngrok.get_tunnels()
        if tunnels:
            return tunnels[0].public_url
    
        # 新規接続
        url = ngrok.connect(8501).public_url
        return url
    except:
        return None

# ngrokトークンの入力欄（セッションステートで管理）
if 'ngrok_token' not in st.session_state:
    st.session_state.ngrok_token = ""

ngrok_token_input = st.sidebar.text_input("ngrok Authtoken (必要な場合)", type="password", key="token_input")

if st.sidebar.button("外部公開を開始する"):
    with st.sidebar.status("接続を試みています..."):
        # トークンがあれば使用
        token_to_use = ngrok_token_input if ngrok_token_input else None
        
        # 接続試行
        public_url = get_ngrok_tunnel(token_to_use)
        
        if public_url:
            st.sidebar.success("接続成功！")
            st.sidebar.write("以下のURLにスマホからアクセスしてください:")
            st.sidebar.code(public_url)
        else:
            st.sidebar.error("接続に失敗しました。")
            st.sidebar.warning(
                "ngrokのAuthtokenが必要な場合があります。\n"
                "1. ngrok.comに登録（無料）\n"
                "2. Your Authtokenをコピー\n"
                "3. 上記の入力欄に貼り付けて再試行してください。"
            )


# ==========================================
# メイン処理
# ==========================================

# 1. 写真アップロード
uploaded_files = st.file_uploader("工事写真をアップロードしてください（複数可）", 
                                  type=['jpg', 'jpeg', 'png'], 
                                  accept_multiple_files=True)

if uploaded_files:
    st.write(f"📸 {len(uploaded_files)} 枚の写真が選択されました。")
    st.markdown("---")

    # データを保持するリスト
    data_list = []

    # 2. 各写真の情報入力
    for i, file in enumerate(uploaded_files):
        col1, col2 = st.columns([1, 2])
        
        # 画像を開く
        image = Image.open(file)
        
        # Exifによる回転補正（スマホ写真で重要）
        try:
            for orientation in ExifTags.TAGS.keys():
                if ExifTags.TAGS[orientation] == 'Orientation':
                    break
            exif = image._getexif()
            if exif is not None:
                orientation = exif.get(orientation)
                if orientation == 3:
                    image = image.rotate(180, expand=True)
                elif orientation == 6:
                    image = image.rotate(270, expand=True)
                elif orientation == 8:
                    image = image.rotate(90, expand=True)
        except (AttributeError, KeyError, IndexError):
            # Exif情報がない、または読み取れない場合は何もしない
            pass

        # 左側：画像プレビュー
        with col1:
            st.image(image, caption=f"写真 {i+1}", use_column_width=True)
        
        # 右側：入力フォーム
        with col2:
            st.subheader(f"写真 {i+1} の情報")
            number = st.text_input(f"番号 (例: ①, {i+1})", value=f"①", key=f"num_{i}")
            content = st.text_area(f"工事箇所・内容", value="トイレ手すり取り付け", key=f"txt_{i}")
            
            # リストに追加（画像オブジェクトもここで一時保存せず、ファイルポインタを使い回すためにリストには入れないほうが安全だが、
            # StreamlitのUploadedFileはseek(0)すれば再読込可能。
            # 回転補正済みの画像を後で使いたいので、処理用には補正済みimageを渡す必要があるが、
            # 現在のループ構造だと後で再処理している。
            # ここではシンプルに入力情報だけ保持し、エクセル作成時にもう一度開くか、または補正済み画像を保持するか。
            # メモリ効率を考えるとファイルから都度読み込むのが良いが、回転補正処理を2回書くことになる。
            # 今回は機能追加箇所で回転補正済み画像を保存する形に変えるのがスマート。
            
            data_list.append({
                "original_file": file, # 元ファイル
                "number": number,
                "content": content,
                "full_text": f"{number} {content}"
            })
        st.markdown("---")

    # お客様名入力
    customer_name = st.sidebar.text_input("お客様名", placeholder="例：山田 太郎 様")

    # 3. エクセル作成ボタン
    if st.button("エクセル台帳を作成する"):
        
        # エクセル作成処理（メモリ上で実行）
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "工事写真台帳"
        
        # A4縦設定
        # A4縦設定
        ws.page_setup.paperSize = 9 # A4
        ws.page_setup.orientation = 'portrait'

        # 列幅の設定（A列とB列をメインに使用）
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 45 # 少し広めに

        # タイトル表示（お客様名）
        title_font = Font(name='Meiryo', size=14, bold=True)
        ws["A1"] = f"{customer_name}　施工前写真" if customer_name else "施工前写真"
        ws["A1"].font = title_font

        # フォント設定
        font_style = Font(name='Meiryo', size=11, bold=True)
        align_style = Alignment(horizontal='left', vertical='top', wrap_text=True)

        current_row = 2 # タイトルがあるので2行目から開始
        col_index = 0 # 0:左, 1:右

        for item in data_list:
            # 画像を再度開き、回転補正を行う（プレビュー時と同じ処理）
            item["original_file"].seek(0)
            img_pil = Image.open(item["original_file"])
            
            # Exif情報の取得用変数
            exif_date = None

            try:
                # Exif取得と回転補正
                exif = img_pil._getexif()
                if exif:
                    # 日付取得 (DateTimeOriginal: 36867)
                    if 36867 in exif:
                        exif_date_str = exif[36867] # "YYYY:MM:DD HH:MM:SS"
                        try:
                            dt = datetime.datetime.strptime(exif_date_str, '%Y:%m:%d %H:%M:%S')
                            exif_date = dt.strftime('%Y.%m.%d')
                        except:
                            pass
                    
                    # 回転補正
                    for orientation in ExifTags.TAGS.keys():
                        if ExifTags.TAGS[orientation] == 'Orientation':
                            break
                    orient = exif.get(orientation)
                    if orient == 3:
                        img_pil = img_pil.rotate(180, expand=True)
                    elif orient == 6:
                        img_pil = img_pil.rotate(270, expand=True)
                    elif orient == 8:
                        img_pil = img_pil.rotate(90, expand=True)
            except:
                pass

            draw = ImageDraw.Draw(img_pil)
            
            # フォントサイズを倍にする (40 -> 80)
            font_size = 80
            try:
                # Windows向けの日本語フォント設定
                font = ImageFont.truetype("C:\\Windows\\Fonts\\meiryo.ttc", font_size)
            except:
                try:
                    font = ImageFont.truetype("DejaVuSans.ttf", font_size)
                except:
                    font = ImageFont.load_default()

            # 日付テキストの決定
            text_to_draw = None
            if date_mode == "指定日を入れる":
                text_to_draw = date_text_fixed
            elif date_mode == "写真の撮影日(Exif)":
                text_to_draw = exif_date # 取得できていれば文字列、なければNone
            # "印字しない" の場合は None のまま

            # 日付を描画（テキストがある場合のみ）
            if text_to_draw:
                text_color = (255, 165, 0) # オレンジ
                width, height = img_pil.size
                
                # 文字幅の計算も考慮して位置調整
                # 簡易計算: 文字数 * フォントサイズの半角換算 * 係数
                text_len = len(text_to_draw) * (font_size / 2) 
                # 右端から少し余裕を持たせる (height - 120 くらいに調整)
                draw.text((width - 100 - text_len, height - 120), text_to_draw, fill=text_color, font=font)

            # エクセルに配置するためにバイトストリームに保存
            img_byte_arr = io.BytesIO()
            img_pil.save(img_byte_arr, format='JPEG')
            img_byte_arr.seek(0)
            
            # エクセル配置用画像オブジェクト作成
            xl_img = ExcelImage(img_byte_arr)
            xl_img.width = 320
            xl_img.height = 240
            
            # セル位置決定
            col_letter = 'A' if col_index == 0 else 'B'
            
            # テキストセル
            cell = ws[f"{col_letter}{current_row}"]
            cell.value = item["full_text"]
            cell.font = font_style
            cell.alignment = align_style
            
            # 画像セル
            img_row = current_row + 1
            ws.add_image(xl_img, f"{col_letter}{img_row}")
            
            # 行の高さ
            ws.row_dimensions[current_row].height = 30
            ws.row_dimensions[img_row].height = 190

            # 次の配置へ
            if col_index == 1:
                col_index = 0
                current_row += 2 # 次の段へ
            else:
                col_index = 1

        # 保存
        wb.save(output)
        output.seek(0)

        # 4. ダウンロードボタン表示
        file_date_suffix = datetime.date.today().strftime('%Y%m%d')
        st.success("作成完了しました！")
        st.download_button(
            label="エクセルファイルをダウンロード",
            data=output,
            file_name=f"工事写真台帳_{file_date_suffix}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
